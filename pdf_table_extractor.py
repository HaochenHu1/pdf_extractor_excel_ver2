from __future__ import annotations

import argparse
import re
import subprocess
import sys
from collections import Counter
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

import pandas as pd
import fitz  #PyMuPDF
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from paragraph_metric_extractor import (
    SectionExtractionResult,
    default_section_configs,
    demo_extract_market_section_metrics,
    extract_configured_sections_from_pdf,
)
from shandong_monthly_extractor import (
    ShandongExtractionResult,
    extract_shandong_market_disclosure_monthly_report,
)

@dataclass
class ExtractedTable:
    df: pd.DataFrame
    page: int
    engine: str
    score: float
    title: Optional[str] = None
    layout_meta: Optional[Dict[str, Any]] = None

#Command line parameters
#`input_pdf` for pdf path
#`-o/--output` for export path
#`--pages` for pages needed to be processed
#`--mode` for selection of models
#`-prefer` only affects Camelot
#`--min-rows`/`--min-cols`/`--min-filled-ratio` for threshold for filtering fake tables
#`--accuracy-threshold` lowest accuracy rate control
#`--ocr-lang` language selection for pdf
#`--borderless` asking script to detect tables without border
#`--verbose` for logging print
def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Extract tables from a PDF and write them to an Excel workbook."
    )
    parser.add_argument(
        "input_path",
        type=Path,
        help="Path to a PDF file or a folder that contains PDF files",
    )
    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        default=None,
        help="Output .xlsx path for single-file input only",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=None,
        help="Output directory for batch/folder input. Defaults to <input_folder>/extracted_tables",
    )
    parser.add_argument(
        "--pages",
        default="all",
        help='Pages to process. Examples: "all", "1", "1,3,5", "2-6"',
    )
    parser.add_argument(
        "--mode",
        choices=["auto", "camelot", "pdfplumber", "img2table"],
        default="camelot",
        help="Extraction backend selection (default: camelot for text-based PDFs)",
    )
    parser.add_argument(
        "--prefer",
        choices=["stream", "lattice", "both"],
        default="both",
        help="Camelot extraction style preference for text-based PDFs",
    )
    parser.add_argument(
        "--min-rows",
        type=int,
        default=2,
        help="Minimum number of rows a table must have after cleanup",
    )
    parser.add_argument(
        "--min-cols",
        type=int,
        default=2,
        help="Minimum number of columns a table must have after cleanup",
    )
    parser.add_argument(
        "--min-filled-ratio",
        type=float,
        default=0.15,
        help="Minimum non-empty cell ratio required to keep a table",
    )
    parser.add_argument(
        "--accuracy-threshold",
        type=float,
        default=50.0,
        help="Minimum Camelot accuracy for keeping a table when available",
    )
    parser.add_argument(
        "--ocr-lang",
        default="eng",
        help="Tesseract OCR language for img2table fallback",
    )
    parser.add_argument(
        "--ocr-lang-auto",
        action="store_true",
        help="Auto-tune OCR language/settings for Chinese scanned PDFs when possible",
    )
    parser.add_argument(
        "--borderless",
        action="store_true",
        help="Enable borderless table extraction for img2table",
    )
    parser.add_argument(
        "--img2table-min-confidence",
        type=int,
        default=50,
        help="img2table minimum OCR confidence (0-99); lower can help noisy scanned Chinese docs",
    )
    parser.add_argument(
        "--verbose",
        action="store_true",
        help="Print extraction progress",
    )
    parser.add_argument(
        "--excel-style-mode",
        choices=["basic", "off"],
        default="basic",
        help="Excel styling mode for merged cells: basic applies centered wrapped labels; off disables extra styling.",
    )
    parser.add_argument(
        "--row-compact",
        dest="row_compact",
        action="store_true",
        help="Compact sparse continuation/header rows by merging them upward.",
    )
    parser.add_argument(
        "--no-row-compact",
        dest="row_compact",
        action="store_false",
        help="Disable sparse row compaction.",
    )
    parser.set_defaults(row_compact=True)
    parser.add_argument(
        "--row-compact-empty-ratio",
        type=float,
        default=0.8,
        help="Minimum empty-cell ratio to consider a row sparse for upward compaction.",
    )
    parser.add_argument(
        "--row-compact-header-rows",
        type=int,
        default=5,
        help="Only compact rows within the first N rows (header zone). Use 0 to disable this limit.",
    )
    parser.add_argument(
        "--demo-section-metrics",
        action="store_true",
        help="Run a built-in demo for paragraph metric extraction and exit.",
    )
    return parser.parse_args()


def collect_input_pdfs(input_path: Path) -> List[Path]:
    if input_path.is_file():
        if input_path.suffix.lower() != ".pdf":
            raise ValueError(f"Input file is not a PDF: {input_path}")
        return [input_path]

    if input_path.is_dir():
        pdfs = sorted(p for p in input_path.iterdir() if p.is_file() and p.suffix.lower() == ".pdf")
        return pdfs

    raise ValueError(f"Input path does not exist: {input_path}")

#Only print log when verbose is true
def log(message: str, verbose: bool = True) -> None:
    if verbose:
        print(message)

#Convert a page number expression provided by the user 
#into an actual list of page numbers
def expand_page_ranges(pages_spec: str, max_pages: int) -> List[int]:
    if pages_spec.lower() == "all":
        return list(range(1, max_pages + 1))

    pages: set[int] = set()
    chunks = [chunk.strip() for chunk in pages_spec.split(",") if chunk.strip()]
    pattern = re.compile(r"^(\d+)(?:-(\d+))?$")

    for chunk in chunks:
        match = pattern.match(chunk)
        if not match:
            raise ValueError(f"Invalid page spec: {chunk!r}")
        start = int(match.group(1))
        end = int(match.group(2) or start)
        if start < 1 or end < 1 or start > end:
            raise ValueError(f"Invalid page range: {chunk!r}")
        if start > max_pages:
            continue
        end = min(end, max_pages)
        pages.update(range(start, end + 1))

    return sorted(pages)

#Cleaning/ Standardizing cell contents
#Converts `None` into an empty string, casts any value to a string, 
#Replaces carriage returns and line breaks with spaces, 
#Uses regex to condense multiple consecutive whitespace characters into a single space, 
#And finally removes leading and trailing whitespace
def normalize_cell(value: object) -> str:
    if value is None:
        return ""
    text = str(value)
    text = text.replace("\r", " ").replace("\n", " ")
    text = re.sub(r"\s+", " ", text).strip()
    if text.lower() in {"nan", "none", "null", "<na>"}:
        return ""
    return text


def normalize_split_numeric_fragments(text: str) -> str:
    normalized = text
    normalized = re.sub(r"(?<!\d)(\d+)\s*\.\s+(\d{1,3})(?!\d)", r"\1.\2", normalized)
    normalized = re.sub(r"(\d+\.)\s+(\d{1,3})\b", r"\1\2", normalized)
    normalized = re.sub(r"(\d+\.\d)\s+(\d)\b", r"\1\2", normalized)
    normalized = re.sub(r"(\d+\.\d{2})\s+([%‰])\b", r"\1\2", normalized)
    normalized = re.sub(r"(\d)\s+(\d{3}(?:\.\d+)?)\b", r"\1\2", normalized)
    normalized = re.sub(r"(\d),\s+(\d{3}\b)", r"\1,\2", normalized)
    return normalized.strip()

#Creates a copy of the original table using `copy()`; 
#Then cleanses every individual cell using `cleaned.map(normalize_cell)`
#Removes any rows and columns that are entirely empty using apply(..., axis=1)`,
#Finally, it resets the index to start from 0 and standardizes the column names to `col_1`
def clean_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    cleaned = df.copy()
    cleaned = cleaned.map(normalize_cell)

    #Drop fully empty rows and columns.
    non_empty_row_mask = cleaned.apply(
        lambda row: any(normalize_cell(v) != "" for v in row), axis=1
    )
    cleaned = cleaned.loc[non_empty_row_mask]

    if cleaned.empty:
        return cleaned

    non_empty_col_mask = [
        any(normalize_cell(v) != "" for v in cleaned[col]) for col in cleaned.columns
    ]
    cleaned = cleaned.loc[:, non_empty_col_mask]

    cleaned = cleaned.reset_index(drop=True)
    cleaned.columns = [f"col_{i+1}" for i in range(cleaned.shape[1])]
    return cleaned


def _looks_like_monthly_attach1_table(df: pd.DataFrame) -> bool:
    if df.empty:
        return False
    header_zone = " ".join(
        normalize_cell(df.iat[r, c]) for r in range(min(6, df.shape[0])) for c in range(df.shape[1])
    )
    compact_zone = re.sub(r"\s+", "", header_zone)
    keywords = ("附表1", "序号", "机组编号", "单机容量", "总运行小时", "发电小时")
    return sum(1 for keyword in keywords if keyword in compact_zone) >= 2


def _extract_attach1_header_rows(df: pd.DataFrame, max_header_rows: int = 5) -> int:
    header_rows = 0
    for row_idx in range(min(max_header_rows, df.shape[0])):
        row_values = [normalize_cell(v) for v in df.iloc[row_idx].tolist()]
        row_text = " ".join(v for v in row_values if v)
        compact_row_text = re.sub(r"\s+", "", row_text)
        if not row_text:
            continue
        non_empty_count = len([v for v in row_values if v])
        numeric_cells = sum(bool(re.fullmatch(r"[-+]?\d+(?:\.\d+)?%?", v)) for v in row_values if v)
        marker_hit = sum(
            1
            for marker in ("序号", "名称", "机组编号", "单机容量", "本月", "同比", "累计", "小时", "次数", "成功率")
            if marker in compact_row_text
        )
        if marker_hit >= 1 and numeric_cells <= max(1, non_empty_count // 2):
            header_rows = row_idx + 1
        else:
            break
    return max(header_rows, 1)


def _rebuild_attach1_column_names(df: pd.DataFrame, header_rows: int) -> List[str]:
    metric_keywords = ["总运行小时", "发电小时", "抽水小时", "发电调相", "抽水调相", "等效可用系数", "启停次数", "启停成功率", "等效强迫停运率"]
    sub_keywords = ["本月", "同比", "累计"]
    header_matrix: List[List[str]] = []
    for row_idx in range(min(header_rows, df.shape[0])):
        row_vals = [normalize_cell(df.iat[row_idx, c]) for c in range(df.shape[1])]
        carry_metric = ""
        for col_idx, cell in enumerate(row_vals):
            if col_idx < 4:
                continue
            compact_cell = re.sub(r"\s+", "", cell)
            metric_hit = next((m for m in metric_keywords if m in compact_cell), "")
            if metric_hit:
                carry_metric = metric_hit
            elif not cell and carry_metric:
                row_vals[col_idx] = carry_metric
        header_matrix.append(row_vals)

    columns: List[str] = []
    for col_idx in range(df.shape[1]):
        header_fragments = [
            header_matrix[r][col_idx] if r < len(header_matrix) else normalize_cell(df.iat[r, col_idx])
            for r in range(min(header_rows, df.shape[0]))
        ]
        header_text = " ".join(fragment for fragment in header_fragments if fragment)
        compact_header_text = re.sub(r"\s+", "", header_text)
        if col_idx == 0:
            columns.append("序号")
            continue
        if col_idx == 1:
            columns.append("名称")
            continue
        if col_idx == 2:
            columns.append("机组编号")
            continue
        if col_idx == 3:
            columns.append("单机容量（万千瓦）")
            continue

        metric = next((m for m in metric_keywords if m in compact_header_text), "")
        sub = next((s for s in sub_keywords if s in compact_header_text), "")
        if metric and sub:
            columns.append(f"{metric}_{sub}")
        elif metric:
            columns.append(metric)
        elif sub:
            columns.append(f"指标_{sub}")
        elif header_text:
            columns.append(header_text)
        else:
            columns.append(f"指标_{col_idx+1}")
    return columns


def postprocess_monthly_attach1_table(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df
    normalized = df.copy().map(lambda value: normalize_split_numeric_fragments(normalize_cell(value)))
    if not _looks_like_monthly_attach1_table(normalized):
        return normalized

    header_rows = _extract_attach1_header_rows(normalized)
    if header_rows > 0 and normalized.shape[0] > header_rows:
        normalized.columns = _rebuild_attach1_column_names(normalized, header_rows)
        normalized = normalized.iloc[header_rows:].reset_index(drop=True)

    left_cols = [column for column in ["序号", "名称", "机组编号", "单机容量（万千瓦）"] if column in normalized.columns]
    metric_cols = [column for column in normalized.columns if column not in left_cols]
    stitch_cols = [column for column in normalized.columns if column not in ["序号", "名称", "机组编号"]]
    rows_to_drop: set[int] = set()
    for row_idx in range(normalized.shape[0] - 1):
        current_left = [normalize_cell(normalized.at[row_idx, c]) for c in left_cols]
        next_left = [normalize_cell(normalized.at[row_idx + 1, c]) for c in left_cols]
        current_left_empty = sum(1 for v in current_left if v == "") >= max(1, len(current_left) - 1)
        next_left_has_id = any(v != "" for v in next_left)
        if not (current_left_empty and next_left_has_id):
            continue
        merged_any = False
        for stitch_col in stitch_cols:
            first = normalize_cell(normalized.at[row_idx, stitch_col])
            second = normalize_cell(normalized.at[row_idx + 1, stitch_col])
            merged = ""
            if first.endswith(".") and re.fullmatch(r"\d{1,3}", second):
                merged = f"{first}{second}"
            elif first and second == "%" and re.fullmatch(r"\d+(?:\.\d+)?", first):
                merged = f"{first}%"
            elif first and not second:
                merged = first
            elif not first and second:
                merged = second
            if merged:
                normalized.at[row_idx + 1, stitch_col] = merged
                merged_any = True
        if merged_any:
            rows_to_drop.add(row_idx)

    if rows_to_drop:
        normalized = normalized.drop(index=sorted(rows_to_drop)).reset_index(drop=True)

    for row_idx in range(normalized.shape[0]):
        if "名称" in normalized.columns and "机组编号" in normalized.columns:
            name_value = normalize_cell(normalized.at[row_idx, "名称"])
            unit_value = normalize_cell(normalized.at[row_idx, "机组编号"])
            if not name_value and re.search(r"[\u4e00-\u9fff]", unit_value):
                normalized.at[row_idx, "名称"] = unit_value
                normalized.at[row_idx, "机组编号"] = ""
            capacity_col_exists = "单机容量（万千瓦）" in normalized.columns
            if (
                capacity_col_exists
                and name_value
                and re.fullmatch(r"\d+(?:\.\d+)?", name_value)
                and not normalize_cell(normalized.at[row_idx, "单机容量（万千瓦）"])
            ):
                normalized.at[row_idx, "单机容量（万千瓦）"] = name_value
                normalized.at[row_idx, "名称"] = ""
        if row_idx > 0 and metric_cols:
            has_metrics = any(normalize_cell(normalized.at[row_idx, column]) for column in metric_cols)
            if has_metrics:
                for left_col in left_cols:
                    if normalize_cell(normalized.at[row_idx, left_col]) == "":
                        normalized.at[row_idx, left_col] = normalize_cell(normalized.at[row_idx - 1, left_col])
    return normalized


def postprocess_tables_for_monthly_report(tables: Sequence[ExtractedTable]) -> List[ExtractedTable]:
    processed: List[ExtractedTable] = []
    for table in tables:
        if table.engine.startswith("pymupdf_") and (table.title or "") == "附表1":
            processed.append(table)
            continue
        processed.append(
            ExtractedTable(
                df=postprocess_monthly_attach1_table(table.df),
                page=table.page,
                engine=table.engine,
                score=table.score,
                title=table.title,
                layout_meta=table.layout_meta,
            )
        )
    return processed


def _compact_row_signature(row: Sequence[object]) -> str:
    return "|".join(re.sub(r"\s+", "", normalize_cell(value)) for value in row)


def _is_header_like_row(row: Sequence[object]) -> bool:
    row_text = _compact_row_signature(row)
    if not row_text:
        return False
    header_markers = ("序号", "名称", "机组编号", "单机容量", "本月", "同比", "累计", "小时", "次数", "成功率")
    marker_hits = sum(1 for marker in header_markers if marker in row_text)
    return marker_hits >= 1


def _attach1_table_title_signal(table: ExtractedTable) -> bool:
    if (table.title or "") == "附表1":
        return True
    return _looks_like_monthly_attach1_table(table.df)


def _table_starts_new_attach_table(df: pd.DataFrame) -> bool:
    if df.empty:
        return False
    head_text = "".join(
        re.sub(r"\s+", "", normalize_cell(df.iat[r, c]))
        for r in range(min(3, df.shape[0]))
        for c in range(df.shape[1])
    )
    return bool(re.search(r"附表[2-9]", head_text))


def _table_starts_new_section_title(df: pd.DataFrame) -> bool:
    if df.empty:
        return False
    head_text = "".join(
        re.sub(r"\s+", "", normalize_cell(df.iat[r, c]))
        for r in range(min(3, df.shape[0]))
        for c in range(df.shape[1])
    )
    return bool(re.search(r"[（(][一二三四五六七八九十]+[）)]", head_text))


def _table_bbox_with_page_size(
    table: ExtractedTable,
) -> Tuple[Optional[Tuple[float, float, float, float]], Optional[float], Optional[float]]:
    if not isinstance(table.layout_meta, dict):
        return None, None, None
    bbox = table.layout_meta.get("table_bbox")
    if not (isinstance(bbox, (list, tuple)) and len(bbox) >= 4):
        return None, None, None
    page_height = table.layout_meta.get("page_height")
    page_width = table.layout_meta.get("page_width")
    page_height_num = float(page_height) if isinstance(page_height, (int, float)) else None
    page_width_num = float(page_width) if isinstance(page_width, (int, float)) else None
    return (float(bbox[0]), float(bbox[1]), float(bbox[2]), float(bbox[3])), page_height_num, page_width_num


def _column_boundary_signature(layout_meta: Optional[Dict[str, Any]]) -> List[float]:
    if not isinstance(layout_meta, dict):
        return []
    canonical_cells = layout_meta.get("cells")
    if not isinstance(canonical_cells, list):
        return []
    rows_info = layout_meta.get("canonical_grid", {})
    col_count = rows_info.get("cols") if isinstance(rows_info, dict) else None
    if not isinstance(col_count, int) or col_count <= 0:
        return []

    per_col: Dict[int, List[Tuple[float, float]]] = {idx: [] for idx in range(col_count)}
    for cell in canonical_cells:
        if not isinstance(cell, dict):
            continue
        col_idx = cell.get("col_idx")
        rel_bbox = cell.get("bbox_rel")
        if not isinstance(col_idx, int) or not isinstance(rel_bbox, (list, tuple)) or len(rel_bbox) < 4:
            continue
        try:
            x0 = float(rel_bbox[0])
            x1 = float(rel_bbox[2])
        except (TypeError, ValueError):
            continue
        per_col.setdefault(col_idx, []).append((x0, x1))

    boundaries: List[float] = []
    for col_idx in range(col_count):
        boxes = per_col.get(col_idx, [])
        if boxes:
            left = sum(v[0] for v in boxes) / len(boxes)
            right = sum(v[1] for v in boxes) / len(boxes)
        else:
            left = float(col_idx) / col_count
            right = float(col_idx + 1) / col_count
        if col_idx == 0:
            boundaries.append(left)
        boundaries.append(right)
    return boundaries


def _column_boundary_similarity(boundaries_a: Sequence[float], boundaries_b: Sequence[float]) -> float:
    if not boundaries_a or not boundaries_b:
        return 0.0
    min_len = min(len(boundaries_a), len(boundaries_b))
    if min_len <= 1:
        return 0.0
    diffs = [abs(float(boundaries_a[idx]) - float(boundaries_b[idx])) for idx in range(min_len)]
    mean_diff = sum(diffs) / min_len
    return max(0.0, 1.0 - mean_diff * 3.5)


def _is_data_like_row(row: Sequence[object]) -> bool:
    values = [normalize_cell(v) for v in row]
    non_empty = [v for v in values if v]
    if not non_empty:
        return False
    text = "".join(non_empty)
    if any(marker in text for marker in ("序号", "名称", "机组编号", "单机容量", "本月", "同比", "累计")):
        return False
    numeric_hits = sum(bool(re.search(r"\d", v)) for v in non_empty)
    return numeric_hits >= max(1, len(non_empty) // 2)


def _extract_serial_numbers(df: pd.DataFrame) -> List[int]:
    if df.empty or df.shape[1] == 0:
        return []
    serials: List[int] = []
    for row_idx in range(df.shape[0]):
        raw = normalize_cell(df.iat[row_idx, 0])
        match = re.search(r"\d{1,4}", raw)
        if not match:
            continue
        try:
            serials.append(int(match.group(0)))
        except ValueError:
            continue
    return serials


def _attach1_continuation_diagnostics(
    prev_table: ExtractedTable, next_table: ExtractedTable
) -> Tuple[bool, Dict[str, Any]]:
    details: Dict[str, Any] = {
        "prev_page": prev_table.page,
        "next_page": next_table.page,
        "signals": 0,
    }
    if next_table.page != prev_table.page + 1:
        details["reason"] = "not_adjacent_pages"
        return False, details
    if _table_starts_new_attach_table(next_table.df) or _table_starts_new_section_title(next_table.df):
        details["reason"] = "new_table_or_section_title"
        return False, details

    signals = 0
    details["col_count_prev"] = prev_table.df.shape[1]
    details["col_count_next"] = next_table.df.shape[1]
    if prev_table.df.shape[1] == next_table.df.shape[1]:
        signals += 2
    elif abs(prev_table.df.shape[1] - next_table.df.shape[1]) <= 1:
        signals += 1

    prev_bbox, prev_page_h, prev_page_w = _table_bbox_with_page_size(prev_table)
    next_bbox, next_page_h, next_page_w = _table_bbox_with_page_size(next_table)
    if (
        prev_bbox
        and next_bbox
        and prev_page_h
        and next_page_h
        and prev_page_w
        and next_page_w
        and prev_page_h > 0
        and next_page_h > 0
        and prev_page_w > 0
        and next_page_w > 0
    ):
        prev_bottom_ratio = prev_bbox[3] / prev_page_h
        next_top_ratio = next_bbox[1] / next_page_h
        if prev_bottom_ratio >= 0.75:
            signals += 1
        if next_top_ratio <= 0.25:
            signals += 1
        prev_left_ratio = prev_bbox[0] / prev_page_w
        next_left_ratio = next_bbox[0] / next_page_w
        width_diff_ratio = abs((prev_bbox[2] - prev_bbox[0]) - (next_bbox[2] - next_bbox[0])) / max(
            1.0, (prev_bbox[2] - prev_bbox[0])
        )
        if abs(prev_left_ratio - next_left_ratio) <= 0.06 and width_diff_ratio <= 0.2:
            signals += 1

    next_head_rows = min(3, next_table.df.shape[0])
    if any(_is_data_like_row(next_table.df.iloc[row_idx].tolist()) for row_idx in range(next_head_rows)):
        signals += 1

    if _attach1_table_title_signal(prev_table) or _attach1_table_title_signal(next_table):
        signals += 1

    prev_serials = _extract_serial_numbers(prev_table.df)
    next_serials = _extract_serial_numbers(next_table.df)
    if prev_serials and next_serials:
        if max(next_serials) >= max(prev_serials) and min(next_serials) <= max(prev_serials) + 3:
            signals += 1

    prev_boundaries = _column_boundary_signature(prev_table.layout_meta)
    next_boundaries = _column_boundary_signature(next_table.layout_meta)
    boundary_similarity = _column_boundary_similarity(prev_boundaries, next_boundaries)
    details["boundary_similarity"] = round(boundary_similarity, 4)
    if boundary_similarity >= 0.78:
        signals += 2
    elif boundary_similarity >= 0.66:
        signals += 1

    details["signals"] = signals
    details["reason"] = "passed" if signals >= 3 else "insufficient_signals"
    return signals >= 3, details


def detect_attach1_continuation(prev_table: ExtractedTable, next_table: ExtractedTable) -> bool:
    ok, _ = _attach1_continuation_diagnostics(prev_table, next_table)
    return ok


def _drop_repeated_attach1_headers(base_df: pd.DataFrame, next_df: pd.DataFrame) -> pd.DataFrame:
    if base_df.empty or next_df.empty:
        return next_df
    base_header_rows = min(max(_extract_attach1_header_rows(base_df), 1), 5, base_df.shape[0])
    candidate_rows = min(base_header_rows, 5, next_df.shape[0])
    drop_count = 0
    for row_idx in range(candidate_rows):
        if not _is_header_like_row(next_df.iloc[row_idx].tolist()):
            break
        base_row = base_df.iloc[min(row_idx, base_header_rows - 1)].tolist()
        next_row = next_df.iloc[row_idx].tolist()
        if _compact_row_signature(base_row) == _compact_row_signature(next_row):
            drop_count = row_idx + 1
            continue
        break
    if drop_count <= 0:
        return next_df
    return next_df.iloc[drop_count:].reset_index(drop=True)


def _merge_split_row_across_pages(base_df: pd.DataFrame, next_df: pd.DataFrame) -> pd.DataFrame:
    if base_df.empty or next_df.empty:
        return next_df
    shared_cols = [col for col in base_df.columns if col in next_df.columns]
    if not shared_cols:
        return next_df
    left_cols = [col for col in ("序号", "名称", "机组编号") if col in shared_cols]
    if not left_cols:
        left_cols = shared_cols[: min(3, len(shared_cols))]

    last_row = base_df.iloc[-1]
    first_row = next_df.iloc[0]
    last_left_values = [normalize_cell(last_row[col]) for col in left_cols]
    first_left_values = [normalize_cell(first_row[col]) for col in left_cols]
    last_left_empty = sum(1 for value in last_left_values if value == "")
    first_left_empty = sum(1 for value in first_left_values if value == "")
    last_trailing_empty = 0
    first_leading_empty = 0
    for col in reversed(shared_cols):
        if normalize_cell(last_row[col]) == "":
            last_trailing_empty += 1
        else:
            break
    for col in shared_cols:
        if normalize_cell(first_row[col]) == "":
            first_leading_empty += 1
        else:
            break

    looks_like_split_row = (
        last_left_empty > 0
        or first_left_empty > 0
        or (last_trailing_empty >= max(2, len(shared_cols) // 4) and first_leading_empty >= 1)
    )
    if not looks_like_split_row:
        return next_df

    merged_any = False
    merged_row = last_row.copy()
    for col in shared_cols:
        last_val = normalize_cell(last_row[col])
        first_val = normalize_cell(first_row[col])
        merged_val = None
        if last_val.endswith(".") and re.fullmatch(r"\d{1,3}", first_val):
            merged_val = f"{last_val}{first_val}"
        elif last_val and first_val == "%" and re.fullmatch(r"\d+(?:\.\d+)?", last_val):
            merged_val = f"{last_val}%"
        elif not last_val and first_val:
            merged_val = first_val
        elif last_val and first_val and last_val != first_val and col in left_cols:
            merged_val = f"{last_val}{first_val}"
        if merged_val is not None:
            merged_row[col] = merged_val
            merged_any = True

    if not merged_any:
        return next_df
    base_df.iloc[-1] = merged_row
    return next_df.iloc[1:].reset_index(drop=True)


def _drop_tail_head_duplicate_rows(base_df: pd.DataFrame, next_df: pd.DataFrame) -> pd.DataFrame:
    if base_df.empty or next_df.empty:
        return next_df
    overlap_limit = min(3, len(base_df), len(next_df))
    if overlap_limit <= 0:
        return next_df
    drop_rows = 0
    for overlap in range(overlap_limit, 0, -1):
        base_slice = base_df.tail(overlap)
        next_slice = next_df.head(overlap)
        same = True
        for row_idx in range(overlap):
            base_row = _compact_row_signature(base_slice.iloc[row_idx].tolist())
            next_row = _compact_row_signature(next_slice.iloc[row_idx].tolist())
            if base_row != next_row:
                same = False
                break
        if same:
            drop_rows = overlap
            break
    if drop_rows <= 0:
        return next_df
    return next_df.iloc[drop_rows:].reset_index(drop=True)


def _clean_text_spacing_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df
    return df.map(lambda value: normalize_split_numeric_fragments(normalize_cell(value)))


def stitch_attach1_across_pages(tables: Sequence[ExtractedTable]) -> List[ExtractedTable]:
    if not tables:
        return []
    sorted_tables = sorted(tables, key=lambda table: (table.page, -table.score))
    first_attach1_idx_opt = next((idx for idx, table in enumerate(sorted_tables) if _attach1_table_title_signal(table)), None)
    if first_attach1_idx_opt is None:
        return []
    first_attach1_idx = int(first_attach1_idx_opt)

    stitched_table = sorted_tables[first_attach1_idx]
    stitched_df = stitched_table.df.copy()
    last_page = stitched_table.page
    previous_table = stitched_table
    for table in sorted_tables[first_attach1_idx + 1 :]:
        if table.page < last_page:
            continue
        if table.page > last_page + 1:
            break
        if not detect_attach1_continuation(previous_table, table):
            break
        next_df = _drop_repeated_attach1_headers(stitched_df, table.df.copy())
        next_df = _merge_split_row_across_pages(stitched_df, next_df)
        next_df = _drop_tail_head_duplicate_rows(stitched_df, next_df)
        if next_df.empty:
            last_page = table.page
            previous_table = table
            continue
        stitched_df = pd.concat([stitched_df, next_df], ignore_index=True)
        last_page = table.page
        previous_table = table

    return [
        ExtractedTable(
            df=_clean_text_spacing_dataframe(stitched_df).reset_index(drop=True),
            page=stitched_table.page,
            engine=stitched_table.engine,
            score=stitched_table.score,
            title="附表1",
            layout_meta=stitched_table.layout_meta,
        )
    ]


def _attach1_keyword_hits(df: pd.DataFrame) -> int:
    if df.empty:
        return 0
    header_text = "".join(
        re.sub(r"\s+", "", normalize_cell(df.iat[r, c]))
        for r in range(min(5, df.shape[0]))
        for c in range(df.shape[1])
    )
    keywords = ("序号", "名称", "机组编号", "单机容量", "总运行小时", "发电小时", "抽水小时", "本月", "累计")
    return sum(1 for keyword in keywords if keyword in header_text)


def extract_attach1_with_border_grid(input_pdf: Path, verbose: bool = False) -> List[ExtractedTable]:
    try:
        document = fitz.open(str(input_pdf))
    except Exception:
        return []

    candidates_by_page: Dict[int, ExtractedTable] = {}
    pending_attach1: Optional[ExtractedTable] = None
    try:
        for page_idx, page in enumerate(document):
            page_text = page.get_text("text") or ""
            page_has_attach1_title = "附表1" in page_text
            if not page_has_attach1_title and pending_attach1 is None and not candidates_by_page:
                continue

            page_candidates: List[ExtractedTable] = []
            for strategy in ("lines_strict", "lines"):
                try:
                    finder = page.find_tables(strategy=strategy)
                except Exception:
                    continue
                tables = getattr(finder, "tables", []) or []
                for table in tables:
                    row_count = int(getattr(table, "row_count", 0))
                    col_count = int(getattr(table, "col_count", 0))
                    min_col_count = 8
                    min_row_count = 8
                    if pending_attach1 is not None:
                        min_col_count = max(6, pending_attach1.df.shape[1] - 2)
                        min_row_count = 2
                    if row_count < min_row_count or col_count < min_col_count:
                        if verbose:
                            print(
                                f"[{input_pdf.name}] p{page_idx+1} skip table rows={row_count} cols={col_count} "
                                f"(need rows>={min_row_count}, cols>={min_col_count})"
                            )
                        continue

                    raw_rows = getattr(table, "rows", None) or []
                    if not raw_rows:
                        continue

                    grid_values: List[List[str]] = []
                    raw_cells: List[Dict[str, object]] = []
                    for row_idx, row_obj in enumerate(raw_rows):
                        row_cells = getattr(row_obj, "cells", None) or []
                        row_values: List[str] = []
                        for col_idx, cell_bbox in enumerate(row_cells):
                            bbox = _to_bbox(cell_bbox)
                            if bbox is None:
                                text = ""
                            else:
                                text = page.get_textbox(fitz.Rect(*bbox)) or ""
                            text = normalize_split_numeric_fragments(normalize_cell(text))
                            row_values.append(text)
                            raw_cells.append(
                                {
                                    "row_idx": row_idx,
                                    "col_idx": col_idx,
                                    "text": text,
                                    "bbox": bbox,
                                }
                            )
                        grid_values.append(row_values)

                    if not grid_values:
                        continue
                    df = pd.DataFrame(grid_values)
                    hits = _attach1_keyword_hits(df)

                    layout_meta = _finalize_layout_meta(raw_cells, df.shape[0], df.shape[1], f"pymupdf_{strategy}")
                    layout_meta["page_width"] = float(page.rect.width)
                    layout_meta["page_height"] = float(page.rect.height)
                    page_candidates.append(
                        ExtractedTable(
                        df=df.map(lambda value: normalize_split_numeric_fragments(normalize_cell(value))),
                        page=page_idx + 1,
                        engine=f"pymupdf_{strategy}",
                        score=float(hits),
                        title="附表1",
                        layout_meta=layout_meta,
                    )
                    )

            if not page_candidates:
                if verbose and pending_attach1 is not None:
                    print(
                        f"[{input_pdf.name}] p{page_idx+1} no candidates; clear pending attach1 from p{pending_attach1.page}"
                    )
                pending_attach1 = None
                continue

            accepted: Optional[ExtractedTable] = None
            strong_candidates = [
                candidate
                for candidate in page_candidates
                if candidate.score >= (4 if page_has_attach1_title else 5) or page_has_attach1_title
            ]
            if strong_candidates:
                accepted = max(strong_candidates, key=lambda item: item.score)
            elif pending_attach1 is not None:
                continuation_candidates: List[Tuple[float, ExtractedTable]] = []
                for candidate in page_candidates:
                    probe_prev = ExtractedTable(
                        df=pending_attach1.df,
                        page=pending_attach1.page,
                        engine=pending_attach1.engine,
                        score=pending_attach1.score,
                        title="附表1",
                        layout_meta=pending_attach1.layout_meta,
                    )
                    probe_next = ExtractedTable(
                        df=candidate.df,
                        page=candidate.page,
                        engine=candidate.engine,
                        score=candidate.score,
                        title=None,
                        layout_meta=candidate.layout_meta,
                    )
                    if _table_starts_new_attach_table(candidate.df) or _table_starts_new_section_title(candidate.df):
                        if verbose:
                            print(f"[{input_pdf.name}] p{page_idx+1} candidate rejected: new title/section signal")
                        continue
                    ok, diag = _attach1_continuation_diagnostics(probe_prev, probe_next)
                    if not ok:
                        if verbose:
                            print(
                                f"[{input_pdf.name}] p{page_idx+1} continuation check failed "
                                f"against pending p{pending_attach1.page}: {diag}"
                            )
                        continue
                    boundary_score = _column_boundary_similarity(
                        _column_boundary_signature(pending_attach1.layout_meta),
                        _column_boundary_signature(candidate.layout_meta),
                    )
                    candidate_score = boundary_score + min(candidate.score / 10.0, 0.5)
                    if verbose:
                        print(
                            f"[{input_pdf.name}] p{page_idx+1} continuation match with pending p{pending_attach1.page}: "
                            f"diag={diag}, rank_score={candidate_score:.3f}"
                        )
                    continuation_candidates.append((candidate_score, candidate))
                if continuation_candidates:
                    accepted = max(continuation_candidates, key=lambda item: item[0])[1]

            if accepted is not None:
                previous = candidates_by_page.get(page_idx + 1)
                if previous is None or accepted.score > previous.score:
                    candidates_by_page[page_idx + 1] = accepted
                if verbose:
                    print(
                        f"[{input_pdf.name}] p{page_idx+1} keep attach1 candidate rows={accepted.df.shape[0]} "
                        f"cols={accepted.df.shape[1]} score={accepted.score:.2f}; pending=ON"
                    )
                pending_attach1 = accepted
            else:
                if verbose and pending_attach1 is not None:
                    print(f"[{input_pdf.name}] p{page_idx+1} no accepted continuation; pending cleared")
                pending_attach1 = None
    finally:
        document.close()

    candidates = [candidates_by_page[page] for page in sorted(candidates_by_page)]
    if candidates and verbose:
        pages_text = ", ".join(str(candidate.page) for candidate in candidates)
        print(f"[{input_pdf.name}] 附表1 border-grid candidate pages={pages_text}")
    return candidates


def select_attach1_tables_for_monthly_report(tables: Sequence[ExtractedTable]) -> List[ExtractedTable]:
    return stitch_attach1_across_pages(tables)


def drop_near_duplicate_columns(df: pd.DataFrame, similarity_threshold: float = 0.95) -> pd.DataFrame:
    if df.empty or df.shape[1] <= 1:
        return df

    kept_columns: List[str] = []
    total_rows = max(df.shape[0], 1)

    for col in df.columns:
        current = df[col].astype(str)
        is_duplicate = False
        for kept_col in kept_columns:
            kept = df[kept_col].astype(str)
            same_ratio = float((current == kept).sum()) / total_rows
            if same_ratio >= similarity_threshold:
                is_duplicate = True
                break
        if not is_duplicate:
            kept_columns.append(col)

    deduped = df.loc[:, kept_columns].copy()
    deduped.columns = [f"col_{i+1}" for i in range(deduped.shape[1])]
    return deduped

#Using a score to determine whether a table is too sparse
def dataframe_filled_ratio(df: pd.DataFrame) -> float:
    if df.empty:
        return 0.0
    total = df.shape[0] * df.shape[1]
    if total == 0:
        return 0.0
    filled = int((df != "").sum().sum())
    return filled / total

#Filter to determine if a table is worth keeping
def looks_like_table(
    df: pd.DataFrame,
    min_rows: int,
    min_cols: int,
    min_filled_ratio: float,) -> bool:
    if df.empty:
        return False
    if df.shape[0] < min_rows or df.shape[1] < min_cols:
        return False
    if dataframe_filled_ratio(df) < min_filled_ratio:
        return False
    return True


def compact_sparse_rows(
    df: pd.DataFrame,
    enabled: bool = True,
    empty_ratio_threshold: float = 0.8,
    header_rows_limit: int = 5,
) -> pd.DataFrame:
    if not enabled or df.empty or df.shape[0] <= 1:
        return df

    threshold = min(0.99, max(0.0, float(empty_ratio_threshold)))
    header_limit = max(0, int(header_rows_limit))

    rows = [list(df.iloc[i].astype(str).map(normalize_cell)) for i in range(df.shape[0])]
    out_rows: List[List[str]] = []

    number_pattern = re.compile(r"^[+-]?\d+(?:[\.,]\d+)?(?:%|[a-zA-Z]+)?$")
    for row_idx, row in enumerate(rows):
        if not out_rows:
            out_rows.append(row)
            continue

        empties = sum(1 for cell in row if cell == "")
        empty_ratio = empties / max(len(row), 1)
        non_empty_cells = [cell for cell in row if cell != ""]
        numeric_cells = sum(1 for cell in non_empty_cells if number_pattern.match(cell.replace(" ", "")))

        within_header_zone = header_limit == 0 or row_idx < header_limit
        should_merge_up = (
            within_header_zone
            and empty_ratio >= threshold
            and numeric_cells <= 1
            and len(non_empty_cells) > 0
        )

        if not should_merge_up:
            out_rows.append(row)
            continue

        prev = out_rows[-1]
        merged = prev[:]
        for col_idx, cell in enumerate(row):
            if cell == "":
                continue
            if merged[col_idx] == "":
                merged[col_idx] = cell
        out_rows[-1] = merged

    compacted = pd.DataFrame(out_rows, columns=df.columns)
    compacted = compacted.reset_index(drop=True)
    return compacted

#Determining if two tables are the same
#because `auto` mode would make a table repeatedly processed by different modes
def dataframe_signature(df: pd.DataFrame) -> Tuple[int, int, Tuple[Tuple[str, ...], ...]]:
    rows: List[Tuple[str, ...]] = []
    for row in df.itertuples(index=False, name=None):
        rows.append(tuple(normalize_cell(v) for v in row))
    return df.shape[0], df.shape[1], tuple(rows)


def _to_bbox(raw_bbox: object) -> Optional[Tuple[float, float, float, float]]:
    if raw_bbox is None:
        return None
    if isinstance(raw_bbox, (list, tuple)) and len(raw_bbox) >= 4:
        vals = raw_bbox[:4]
    else:
        return None
    try:
        x0, y0, x1, y1 = (float(vals[0]), float(vals[1]), float(vals[2]), float(vals[3]))
    except (TypeError, ValueError):
        return None
    left, right = sorted((x0, x1))
    top, bottom = sorted((y0, y1))
    return left, top, right, bottom


def _finalize_layout_meta(
    raw_cells: Sequence[Dict[str, object]],
    row_count: int,
    col_count: int,
    source_engine: str,
) -> Dict[str, object]:
    table_bbox: Optional[Tuple[float, float, float, float]] = None
    bboxes = [c["bbox"] for c in raw_cells if c.get("bbox") is not None]
    if bboxes:
        xs0 = [float(b[0]) for b in bboxes]
        ys0 = [float(b[1]) for b in bboxes]
        xs1 = [float(b[2]) for b in bboxes]
        ys1 = [float(b[3]) for b in bboxes]
        table_bbox = (min(xs0), min(ys0), max(xs1), max(ys1))

    table_w = (table_bbox[2] - table_bbox[0]) if table_bbox else 0.0
    table_h = (table_bbox[3] - table_bbox[1]) if table_bbox else 0.0

    canonical_cells: List[Dict[str, object]] = []
    for cell in raw_cells:
        row_idx = int(cell["row_idx"])
        col_idx = int(cell["col_idx"])
        if row_idx < 0 or col_idx < 0 or row_idx >= row_count or col_idx >= col_count:
            continue

        bbox = cell.get("bbox")
        rel_bbox = None
        if table_bbox is not None and bbox is not None and table_w > 0 and table_h > 0:
            rel_bbox = (
                (float(bbox[0]) - table_bbox[0]) / table_w,
                (float(bbox[1]) - table_bbox[1]) / table_h,
                (float(bbox[2]) - table_bbox[0]) / table_w,
                (float(bbox[3]) - table_bbox[1]) / table_h,
            )

        canonical_cells.append(
            {
                "row_idx": row_idx,
                "col_idx": col_idx,
                "text": normalize_cell(cell.get("text", "")),
                "bbox": bbox,
                "bbox_rel": rel_bbox,
            }
        )

    return {
        "source_engine": source_engine,
        "table_bbox": table_bbox,
        "cells": canonical_cells,
        "canonical_grid": {"rows": row_count, "cols": col_count},
    }


def _default_layout_meta(df: pd.DataFrame, source_engine: str) -> Dict[str, object]:
    raw_cells: List[Dict[str, object]] = []
    for row_idx in range(df.shape[0]):
        for col_idx in range(df.shape[1]):
            raw_cells.append(
                {
                    "row_idx": row_idx,
                    "col_idx": col_idx,
                    "text": normalize_cell(df.iat[row_idx, col_idx]),
                    "bbox": None,
                }
            )
    return _finalize_layout_meta(raw_cells, df.shape[0], df.shape[1], source_engine)


def _collect_camelot_layout_meta(table: object, df: pd.DataFrame, source_engine: str) -> Dict[str, object]:
    raw_cells: List[Dict[str, object]] = []
    table_cells = getattr(table, "cells", None)
    if table_cells:
        for row_idx, row in enumerate(table_cells):
            for col_idx, cell in enumerate(row):
                bbox = _to_bbox(
                    (
                        getattr(cell, "x1", None),
                        getattr(cell, "y1", None),
                        getattr(cell, "x2", None),
                        getattr(cell, "y2", None),
                    )
                )
                text = ""
                if row_idx < df.shape[0] and col_idx < df.shape[1]:
                    text = normalize_cell(df.iat[row_idx, col_idx])
                raw_cells.append(
                    {"row_idx": row_idx, "col_idx": col_idx, "text": text, "bbox": bbox}
                )
    if not raw_cells:
        return _default_layout_meta(df, source_engine)
    return _finalize_layout_meta(raw_cells, df.shape[0], df.shape[1], source_engine)


def _collect_pdfplumber_layout_meta(table_obj: object, df: pd.DataFrame, source_engine: str) -> Dict[str, object]:
    raw_cells: List[Dict[str, object]] = []
    rows = getattr(table_obj, "rows", None) or []
    for row_idx, row in enumerate(rows):
        cells = getattr(row, "cells", None) or []
        for col_idx, bbox_candidate in enumerate(cells):
            bbox = _to_bbox(bbox_candidate)
            text = ""
            if row_idx < df.shape[0] and col_idx < df.shape[1]:
                text = normalize_cell(df.iat[row_idx, col_idx])
            raw_cells.append(
                {"row_idx": row_idx, "col_idx": col_idx, "text": text, "bbox": bbox}
            )

    if not raw_cells:
        return _default_layout_meta(df, source_engine)
    return _finalize_layout_meta(raw_cells, df.shape[0], df.shape[1], source_engine)


def _collect_img2table_layout_meta(table: object, df: pd.DataFrame, source_engine: str) -> Dict[str, object]:
    raw_cells: List[Dict[str, object]] = []

    content = getattr(table, "content", None)
    if isinstance(content, dict):
        for row_idx, row_cells in content.items():
            if not isinstance(row_cells, dict):
                continue
            for col_idx, cell in row_cells.items():
                bbox = _to_bbox(getattr(cell, "bbox", None))
                text = normalize_cell(getattr(cell, "value", ""))
                raw_cells.append(
                    {
                        "row_idx": int(row_idx),
                        "col_idx": int(col_idx),
                        "text": text,
                        "bbox": bbox,
                    }
                )

    if not raw_cells:
        return _default_layout_meta(df, source_engine)
    return _finalize_layout_meta(raw_cells, df.shape[0], df.shape[1], source_engine)


def infer_merged_regions(extracted_table: ExtractedTable) -> List[Dict[str, object]]:
    df = extracted_table.df
    layout_meta = extracted_table.layout_meta or {}
    cells = layout_meta.get("cells", []) if isinstance(layout_meta, dict) else []
    cell_lookup: Dict[Tuple[int, int], Dict[str, object]] = {}
    for cell in cells:
        key = (int(cell.get("row_idx", -1)), int(cell.get("col_idx", -1)))
        cell_lookup[key] = cell

    def bbox_alignment_score(anchor: Optional[Tuple[float, float, float, float]], candidate: Optional[Tuple[float, float, float, float]], axis: str) -> float:
        if anchor is None or candidate is None:
            return 0.5
        if axis == "horizontal":
            top = max(float(anchor[1]), float(candidate[1]))
            bottom = min(float(anchor[3]), float(candidate[3]))
            union_top = min(float(anchor[1]), float(candidate[1]))
            union_bottom = max(float(anchor[3]), float(candidate[3]))
            overlap = max(0.0, bottom - top)
            union = max(1e-6, union_bottom - union_top)
            return min(1.0, overlap / union)
        left = max(float(anchor[0]), float(candidate[0]))
        right = min(float(anchor[2]), float(candidate[2]))
        union_left = min(float(anchor[0]), float(candidate[0]))
        union_right = max(float(anchor[2]), float(candidate[2]))
        overlap = max(0.0, right - left)
        union = max(1e-6, union_right - union_left)
        return min(1.0, overlap / union)

    merged: List[Dict[str, object]] = []
    occupied: set[Tuple[int, int, int, int]] = set()
    for row_idx in range(df.shape[0]):
        for col_idx in range(df.shape[1]):
            anchor_text = normalize_cell(df.iat[row_idx, col_idx])
            if anchor_text == "":
                continue

            anchor_bbox = None
            if (row_idx, col_idx) in cell_lookup:
                anchor_bbox = cell_lookup[(row_idx, col_idx)].get("bbox")

            horiz_end = col_idx
            horiz_geom_scores: List[float] = []
            for next_col in range(col_idx + 1, df.shape[1]):
                next_text = normalize_cell(df.iat[row_idx, next_col])
                if next_text != "":
                    break
                next_bbox = cell_lookup.get((row_idx, next_col), {}).get("bbox")
                horiz_geom_scores.append(bbox_alignment_score(anchor_bbox, next_bbox, "horizontal"))
                horiz_end = next_col
            horiz_span = horiz_end - col_idx + 1

            vert_end = row_idx
            vert_geom_scores: List[float] = []
            for next_row in range(row_idx + 1, df.shape[0]):
                next_text = normalize_cell(df.iat[next_row, col_idx])
                if next_text != "":
                    break
                next_bbox = cell_lookup.get((next_row, col_idx), {}).get("bbox")
                vert_geom_scores.append(bbox_alignment_score(anchor_bbox, next_bbox, "vertical"))
                vert_end = next_row
            vert_span = vert_end - row_idx + 1

            if horiz_span <= 1 and vert_span <= 1:
                continue

            use_horizontal = horiz_span >= vert_span
            if use_horizontal:
                key = (row_idx, col_idx, row_idx, horiz_end)
                if key in occupied:
                    continue
                mean_geom = sum(horiz_geom_scores) / len(horiz_geom_scores) if horiz_geom_scores else 0.5
                confidence = round(min(1.0, 0.45 + 0.55 * mean_geom), 4)
                occupied.add(key)
                merged.append(
                    {
                        "start_row": row_idx,
                        "end_row": row_idx,
                        "start_col": col_idx,
                        "end_col": horiz_end,
                        "confidence": confidence,
                        "method": "geometry+empty_neighbor" if horiz_geom_scores else "empty_neighbor",
                    }
                )
            else:
                key = (row_idx, col_idx, vert_end, col_idx)
                if key in occupied:
                    continue
                mean_geom = sum(vert_geom_scores) / len(vert_geom_scores) if vert_geom_scores else 0.5
                confidence = round(min(1.0, 0.45 + 0.55 * mean_geom), 4)
                occupied.add(key)
                merged.append(
                    {
                        "start_row": row_idx,
                        "end_row": vert_end,
                        "start_col": col_idx,
                        "end_col": col_idx,
                        "confidence": confidence,
                        "method": "geometry+empty_neighbor" if vert_geom_scores else "empty_neighbor",
                    }
                )

    return merged


def sanitize_merge_reject_reason(reason: object, max_len: int = 40) -> str:
    text = str(reason or "").lower()
    text = re.sub(r"[\x00-\x1f\x7f]+", " ", text)
    text = re.sub(r"\s+", "_", text.strip())
    text = re.sub(r"[^a-z0-9_]+", "_", text)
    text = re.sub(r"_+", "_", text).strip("_")
    if not text:
        return "unknown"
    return text[:max_len]


def format_merge_reject_top_reasons(raw_reasons: object, max_distinct: int = 5) -> str:
    if max_distinct <= 0:
        return ""

    flattened: List[object] = []
    if isinstance(raw_reasons, str):
        flattened.extend(part for part in raw_reasons.split(",") if part.strip())
    elif isinstance(raw_reasons, (list, tuple, set)):
        for item in raw_reasons:
            if isinstance(item, dict):
                flattened.append(item.get("reason", ""))
            else:
                flattened.append(item)
    elif isinstance(raw_reasons, dict):
        flattened.append(raw_reasons.get("reason", ""))

    if not flattened:
        return ""

    counts: Counter[str] = Counter()
    for item in flattened:
        token = sanitize_merge_reject_reason(item)
        if token:
            counts[token] += 1

    if not counts:
        return ""

    ranked = sorted(counts.items(), key=lambda pair: (-pair[1], pair[0]))[:max_distinct]
    return "|".join(f"{label}:{count}" for label, count in ranked)


def deduplicate_tables(tables: Sequence[ExtractedTable]) -> List[ExtractedTable]:
    best_by_signature: Dict[Tuple[int, int, Tuple[Tuple[str, ...], ...]], ExtractedTable] = {}
    for table in tables:
        sig = dataframe_signature(table.df)
        current = best_by_signature.get(sig)
        if current is None or table.score > current.score:
            best_by_signature[sig] = table
    deduped = list(best_by_signature.values())
    deduped.sort(key=lambda t: (t.page, t.engine, -t.score))
    return deduped


ENGINE_PRIORITY: Dict[str, int] = {
    "pdfplumber_s1": 0,
    "camelot_lattice": 1,
    "camelot_stream": 2,
    "pdfplumber_s2": 3,
    "img2table": 4,
}


def select_best_table_per_page(tables: Sequence[ExtractedTable], verbose: bool = False) -> List[ExtractedTable]:
    if not tables:
        return []

    def sort_key(table: ExtractedTable) -> Tuple[int, float, int, int, int, str]:
        rows, cols = table.df.shape
        area = rows * cols
        engine_priority = ENGINE_PRIORITY.get(table.engine, 999)
        return (
            table.page,
            -float(table.score),
            -area,
            -rows,
            engine_priority,
            table.engine,
        )

    sorted_tables = sorted(tables, key=sort_key)
    best_by_page: Dict[int, ExtractedTable] = {}
    dropped_counts: Counter[int] = Counter()

    for table in sorted_tables:
        if table.page not in best_by_page:
            best_by_page[table.page] = table
        else:
            dropped_counts[table.page] += 1

    selected = [best_by_page[page] for page in sorted(best_by_page)]

    if verbose:
        for page in sorted(best_by_page):
            kept = best_by_page[page]
            removed = dropped_counts[page]
            if removed > 0:
                log(
                    (
                        f"Page {page}: kept {kept.engine} (score={kept.score:.4f}); "
                        f"removed {removed} alternative table(s)."
                    ),
                    verbose,
                )
        log(
            f"Selected {len(selected)} table(s) from {len(tables)} candidate(s) using best-per-page filtering.",
            verbose,
        )

    return selected

#Determining if a pdf is scanned or text-based
def detect_pdf_kind(input_pdf: Path, sample_pages: int = 3) -> str:
    doc = fitz.open(input_pdf)
    total_chars = 0
    checked = min(sample_pages, len(doc))
    for idx in range(checked):
        text = doc[idx].get_text("text") or ""
        total_chars += len(text.strip())
    doc.close()
    if total_chars >= 40:
        return "text"
    return "scanned"

#`extract_with_camelot(...)` is the first function that actually performs table extraction
#Its purpose is to handle text based PDFs
#It first tries to import camelot. If camelot is not installed, it simply returns
#an empty list so that the whole program does not crash
#Next, it converts the page number list into the string format required by Camelot,
#such as "1,2,3"
#If prefer == "both", it will try two extraction modes in sequence:
#1) stream: relies more on text layout
#2) lattice: relies more on table borders and ruling lines
#For each mode, it calls `camelot.read_pdf(**kwargs)` to extract tables
#After extraction, each table goes through the following steps:
#1) Clean the table content with `clean_dataframe()`
#2) Read metadata such as accuracy and page number from parsing_report
#3) Discard the result if the accuracy is too low or the table content is invalid
#4) Otherwise, convert it into an ExtractedTable object
#The scoring logic uses accuracy / 100 to produce the confidence score
#In summary, if Camelot can successfully read the table, this function prefers to keep
#the better structured result along with the corresponding parsing report
def extract_with_camelot(
    input_pdf: Path,
    pages: List[int],
    prefer: str,
    accuracy_threshold: float,
    min_rows: int,
    min_cols: int,
    min_filled_ratio: float,
    row_compact: bool,
    row_compact_empty_ratio: float,
    row_compact_header_rows: int,
    verbose: bool,) -> List[ExtractedTable]:
    try:
        import camelot
    except ImportError:
        return []

    page_spec = ",".join(str(p) for p in pages)
    flavors = ["stream", "lattice"] if prefer == "both" else [prefer]
    extracted: List[ExtractedTable] = []

    for flavor in flavors:
        try:
            log(f"Trying Camelot ({flavor}) on pages {page_spec}", verbose)
            kwargs = {
                "filepath": str(input_pdf),
                "pages": page_spec,
                "flavor": flavor,
                "suppress_stdout": True,
            }
            if flavor == "stream":
                kwargs.update({"row_tol": 10})
            tables = camelot.read_pdf(**kwargs)
        except Exception as exc:
            log(f"Camelot ({flavor}) failed: {exc}", verbose)
            continue

        for idx, table in enumerate(tables):
            try:
                df = clean_dataframe(table.df)
                df = compact_sparse_rows(
                    df,
                    enabled=row_compact,
                    empty_ratio_threshold=row_compact_empty_ratio,
                    header_rows_limit=row_compact_header_rows,
                )
                report = getattr(table, "parsing_report", {}) or {}
                accuracy = float(report.get("accuracy", 100.0))
                page_num = int(report.get("page", pages[0] if pages else 1))
                score = accuracy / 100.0
                if accuracy < accuracy_threshold:
                    continue
                if not looks_like_table(df, min_rows, min_cols, min_filled_ratio):
                    continue
                extracted.append(
                    ExtractedTable(
                        df=df,
                        page=page_num,
                        engine=f"camelot_{flavor}",
                        score=score,
                        title=f"Camelot {flavor} table {idx + 1}",
                        layout_meta=_collect_camelot_layout_meta(table, df, f"camelot_{flavor}"),
                    )
                )
            except Exception as exc:
                log(f"Skipping Camelot table due to error: {exc}", verbose)
    return extracted

#Defines two strategies for pdfplumber extraction
#1) based on boarder to locate table
#2) based on text allocation to infer table
PDFPLUMBER_SETTINGS: List[Dict[str, object]] = [
    {
        "vertical_strategy": "lines",
        "horizontal_strategy": "lines",
        "intersection_tolerance": 5,},
    {
        "vertical_strategy": "text",
        "horizontal_strategy": "text",
        "min_words_vertical": 2,
        "min_words_horizontal": 1,},
]

#This is the second extraction function
#Begins by attempting to import `pdfplumber`; if the import fails, it returns an empty list
#Subsequently, it opens the PDF and iterates through it page by page, 
#applying the two sets of `PDFPLUMBER_SETTINGS` mentioned above to each page
#Each extracted raw table is first converted into a `pd.DataFrame(raw_table)`, 
#then processed by `clean_dataframe()`, and finally filtered using `looks_like_table()`
def extract_with_pdfplumber(
    input_pdf: Path,
    pages: List[int],
    min_rows: int,
    min_cols: int,
    min_filled_ratio: float,
    row_compact: bool,
    row_compact_empty_ratio: float,
    row_compact_header_rows: int,
    verbose: bool,) -> List[ExtractedTable]:
    try:
        import pdfplumber
    except ImportError:
        return []

    extracted: List[ExtractedTable] = []
    with pdfplumber.open(str(input_pdf)) as pdf:
        for page_num in pages:
            page = pdf.pages[page_num - 1]
            for setting_idx, table_settings in enumerate(PDFPLUMBER_SETTINGS, start=1):
                try:
                    raw_tables = page.find_tables(table_settings=table_settings)
                except Exception as exc:
                    log(
                        f"pdfplumber failed on page {page_num} with setting {setting_idx}: {exc}",
                        verbose,
                    )
                    continue
                for idx, table_obj in enumerate(raw_tables):
                    try:
                        raw_table = table_obj.extract()
                        df = clean_dataframe(pd.DataFrame(raw_table))
                        df = compact_sparse_rows(
                            df,
                            enabled=row_compact,
                            empty_ratio_threshold=row_compact_empty_ratio,
                            header_rows_limit=row_compact_header_rows,
                        )
                        if not looks_like_table(df, min_rows, min_cols, min_filled_ratio):
                            continue
                        score = dataframe_filled_ratio(df)
                        extracted.append(
                            ExtractedTable(
                                df=df,
                                page=page_num,
                                engine=f"pdfplumber_s{setting_idx}",
                                score=score,
                                title=f"pdfplumber setting {setting_idx} table {idx + 1}",
                                layout_meta=_collect_pdfplumber_layout_meta(
                                    table_obj, df, f"pdfplumber_s{setting_idx}"
                                ),
                            )
                        )
                    except Exception as exc:
                        log(f"Skipping pdfplumber table due to error: {exc}", verbose)
    return extracted

#This is the third extraction function
#It first imports the PDF wrapper from img2table.document. If that import fails,
#the function prints a warning and returns an empty list
#Then it tries to initialize TesseractOCR. If OCR setup succeeds, it creates an OCR
#engine with the requested language and a triple thread
#Next, it opens the PDF with `Img2TablePDF(...)`. The page numbers are shifted by -1
#because this library uses zero based page indexing
#The actual extraction happens in pdf.extract_tables(...). At this stage:
#1) borderless_tables controls whether borderless tables should also be detected
#2) min_confidence sets the minimum confidence threshold for accepted results
#After tables are extracted, the function loops through the results, converts each
#table object into a DataFrame, filters out poor quality outputs, computes a
#confidence based score, and wraps valid results as ExtractedTable objects
#In short, this function is the OCR oriented fallback: even when the PDF itself does
#not contain readable text, it still tries to recover table structure from the page image
def extract_with_img2table(
    input_pdf: Path,
    pages: List[int],
    ocr_lang: str,
    borderless: bool,
    min_confidence: int,
    implicit_rows: bool,
    implicit_columns: bool,
    min_rows: int,
    min_cols: int,
    min_filled_ratio: float,
    row_compact: bool,
    row_compact_empty_ratio: float,
    row_compact_header_rows: int,
    verbose: bool,) -> List[ExtractedTable]:
    try:
        from img2table.document import PDF as Img2TablePDF
    except ImportError:
        log("img2table is not installed. Skipping OCR fallback.", verbose)
        return []

    ocr = None
    try:
        from img2table.ocr import TesseractOCR

        ocr = TesseractOCR(n_threads=3, lang=ocr_lang)
    except Exception as exc:
        log(f"Tesseract OCR is unavailable: {exc}", verbose)

    try:
        # For scanned PDFs, relying only on embedded PDF text can miss tables completely.
        # If OCR is available, force image-based extraction to improve recall.
        pdf = Img2TablePDF(
            src=str(input_pdf),
            pages=[p - 1 for p in pages],
            pdf_text_extraction=ocr is None,
        )
        tables_by_page = pdf.extract_tables(
            ocr=ocr,
            implicit_rows=implicit_rows,
            implicit_columns=implicit_columns,
            borderless_tables=borderless,
            min_confidence=min_confidence,
        )
    except Exception as exc:
        log(f"img2table failed: {exc}", verbose)
        return []

    extracted: List[ExtractedTable] = []
    for zero_based_page, tables in tables_by_page.items():
        page_num = zero_based_page + 1
        for idx, table in enumerate(tables):
            try:
                raw_df = getattr(table, "df", None)
                if raw_df is None:
                    continue
                df = clean_dataframe(pd.DataFrame(raw_df))
                df = drop_near_duplicate_columns(df)
                df = compact_sparse_rows(
                    df,
                    enabled=row_compact,
                    empty_ratio_threshold=row_compact_empty_ratio,
                    header_rows_limit=row_compact_header_rows,
                )
                if not looks_like_table(df, min_rows, min_cols, min_filled_ratio):
                    continue
                score = dataframe_filled_ratio(df)
                extracted.append(
                    ExtractedTable(
                        df=df,
                        page=page_num,
                        engine="img2table",
                        score=score,
                        title=f"img2table table {idx + 1}",
                        layout_meta=_collect_img2table_layout_meta(table, df, "img2table"),
                    )
                )
            except Exception as exc:
                log(f"Skipping img2table table due to error: {exc}", verbose)
    return extracted


def tune_ocr_options(ocr_lang: str, borderless: bool, min_confidence: int, auto_tune: bool) -> Tuple[str, bool, int, bool, bool]:
    tuned_lang = ocr_lang
    tuned_borderless = borderless
    tuned_confidence = max(0, min(99, int(min_confidence)))
    implicit_rows = False
    implicit_columns = False

    if not auto_tune:
        return tuned_lang, tuned_borderless, tuned_confidence, implicit_rows, implicit_columns

    lang_lower = tuned_lang.lower()
    has_chinese = "chi" in lang_lower or "zh" in lang_lower
    if has_chinese:
        # Chinese scanned tables often need borderless + implicit structure inference.
        tuned_borderless = True
        tuned_confidence = min(tuned_confidence, 35)
        implicit_rows = True
        implicit_columns = True
        if "+" not in tuned_lang and "eng" not in lang_lower:
            tuned_lang = f"{tuned_lang}+eng"

    return tuned_lang, tuned_borderless, tuned_confidence, implicit_rows, implicit_columns


def get_available_tesseract_languages() -> set[str]:
    try:
        result = subprocess.run(
            ["tesseract", "--list-langs"],
            check=False,
            capture_output=True,
            text=True,
        )
    except Exception:
        return set()

    if result.returncode != 0:
        return set()

    langs: set[str] = set()
    for line in result.stdout.splitlines():
        lang = line.strip()
        if not lang or lang.startswith("List of available languages"):
            continue
        langs.add(lang)
    return langs

#Exports all extracted tables into a single Excel file
#It first makes sure the output folder exists, then opens an Excel writer and saves
#each table to its own worksheet, using names like Table_001, Table_002, and so on.
#It also builds a small summary table that records useful
#metadata for each result, such as page number, table index, score, shape, and title.
#After all sheets are created, the function uses openpyxl to do some light formatting,
#including adjusting column widths and freezing the top row so the file is easier to read.
def _unique_sheet_name(workbook: object, base_name: str) -> str:
    cleaned = base_name.strip() or "SectionMetrics"
    cleaned = cleaned[:31]
    if cleaned not in workbook.sheetnames:
        return cleaned
    for idx in range(1, 200):
        suffix = f"_{idx}"
        candidate = f"{cleaned[:31 - len(suffix)]}{suffix}"
        if candidate not in workbook.sheetnames:
            return candidate
    return f"Section_{len(workbook.sheetnames) + 1}"[:31]


def remove_all_whitespace_in_strings(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df
    cleaned = df.copy()
    cleaned = cleaned.map(
        lambda value: (
            re.sub(r"[\s\u00a0\u3000]+", "", value)
            if isinstance(value, str)
            else value
        )
    )
    return cleaned


def write_excel(
    output_path: Path,
    tables: Sequence[ExtractedTable],
    excel_style_mode: str = "basic",
    section_results: Optional[Sequence[SectionExtractionResult]] = None,
    include_summary_sheet: bool = True,
    table_sheet_base_name: str = "Table",
    table_write_header: bool = True,
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        summary_rows = []
        for idx, table in enumerate(tables, start=1):
            sheet_name = f"{table_sheet_base_name}_{idx:03d}" if table_sheet_base_name != "附表1" else "附表1"
            final_df = remove_all_whitespace_in_strings(table.df)
            final_df.to_excel(writer, index=False, header=table_write_header, sheet_name=sheet_name)
            merged_regions = infer_merged_regions(table)
            merge_conf_avg = (
                round(sum(float(m["confidence"]) for m in merged_regions) / len(merged_regions), 4)
                if merged_regions
                else 0.0
            )
            merge_methods = sorted({str(m["method"]) for m in merged_regions}) if merged_regions else []
            raw_merge_reject_reasons = (
                table.layout_meta.get("merge_reject_reasons", []) if isinstance(table.layout_meta, dict) else []
            )
            summary_rows.append(
                {
                    "sheet_name": sheet_name,
                    "page": table.page,
                    "engine": table.engine,
                    "score": round(table.score, 4),
                    "rows": table.df.shape[0],
                    "cols": table.df.shape[1],
                    "title": table.title or "",
                    "merge_count": len(merged_regions),
                    "merge_confidence_avg": merge_conf_avg,
                    "merge_method": ",".join(merge_methods),
                    "merge_reject_top_reasons": format_merge_reject_top_reasons(raw_merge_reject_reasons),
                }
            )

        if include_summary_sheet:
            summary_df = pd.DataFrame(summary_rows)
            summary_df.to_excel(writer, index=False, sheet_name="_summary")

        workbook = writer.book
        if section_results:
            section_sheet_name = _unique_sheet_name(workbook, "市场情况")
            day_ahead = next((s for s in section_results if s.section_title == "（二）日前市场情况"), None)
            real_time = next((s for s in section_results if s.section_title == "（三）实时市场情况"), None)

            if day_ahead:
                day_ahead_df = pd.DataFrame(
                    day_ahead.rows,
                    columns=["title_month_date", "metric_name", "metric_value", "metric_unit", "report_date"],
                )[["title_month_date", "metric_name", "metric_value", "metric_unit"]]
                day_ahead_df = remove_all_whitespace_in_strings(day_ahead_df)
                day_ahead_df.to_excel(
                    writer,
                    index=False,
                    header=False,
                    sheet_name=section_sheet_name,
                    startrow=1,
                    startcol=0,
                )
            if real_time:
                real_time_df = pd.DataFrame(
                    real_time.rows,
                    columns=["title_month_date", "metric_name", "metric_value", "metric_unit", "report_date"],
                )[["title_month_date", "metric_name", "metric_value", "metric_unit"]]
                real_time_df = remove_all_whitespace_in_strings(real_time_df)
                real_time_df.to_excel(
                    writer,
                    index=False,
                    header=False,
                    sheet_name=section_sheet_name,
                    startrow=1,
                    startcol=5,
                )

            if day_ahead or real_time:
                sheet = workbook[section_sheet_name]
                sheet["A1"] = "（二）日前市场情况"
                sheet["F1"] = "（三）实时市场情况"
                if day_ahead:
                    for row_idx in range(2, 2 + len(day_ahead.rows)):
                        cell = sheet.cell(row=row_idx, column=1)
                        if cell.value is not None:
                            cell.number_format = "yyyy-mm"
                if real_time:
                    for row_idx in range(2, 2 + len(real_time.rows)):
                        cell = sheet.cell(row=row_idx, column=6)
                        if cell.value is not None:
                            cell.number_format = "yyyy-mm"

        for idx, table in enumerate(tables, start=1):
            sheet_name = f"{table_sheet_base_name}_{idx:03d}" if table_sheet_base_name != "附表1" else "附表1"
            sheet = workbook[sheet_name]
            row_offset = 2 if table_write_header else 1
            for region in infer_merged_regions(table):
                start_row = int(region["start_row"]) + row_offset
                end_row = int(region["end_row"]) + row_offset
                start_col = int(region["start_col"]) + 1
                end_col = int(region["end_col"]) + 1
                if start_row == end_row and start_col == end_col:
                    continue
                sheet.merge_cells(
                    start_row=start_row,
                    start_column=start_col,
                    end_row=end_row,
                    end_column=end_col,
                )
                if excel_style_mode == "basic":
                    anchor = sheet.cell(row=start_row, column=start_col)
                    anchor.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for sheet in workbook.worksheets:
            for col_idx, column_cells in enumerate(sheet.columns, start=1):
                values = [str(cell.value) if cell.value is not None else "" for cell in column_cells]
                max_len = max((len(v) for v in values), default=10)
                sheet.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 40)
            sheet.freeze_panes = "A2"


def extract_tables_for_pdf(input_pdf: Path, args: argparse.Namespace) -> List[ExtractedTable]:
    if not input_pdf.exists():
        raise FileNotFoundError(f"Input PDF not found: {input_pdf}")

    try:
        doc = fitz.open(str(input_pdf))
        max_pages = len(doc)
        doc.close()
        pages = expand_page_ranges(args.pages, max_pages)
    except Exception as exc:
        raise RuntimeError(f"Failed to read PDF metadata for {input_pdf}: {exc}") from exc

    if not pages:
        raise ValueError(f"No valid pages selected for {input_pdf}.")

    pdf_kind = detect_pdf_kind(input_pdf)
    log(f"[{input_pdf.name}] Detected PDF type: {pdf_kind}", args.verbose)
    ocr_auto_tune = args.ocr_lang_auto

    ocr_lang = args.ocr_lang
    if not ocr_auto_tune and args.mode == "auto" and pdf_kind == "scanned" and ocr_lang == "eng":
        available_langs = get_available_tesseract_languages()
        if "chi_sim" in available_langs and "eng" in available_langs:
            # For scanned auto mode, bilingual OCR is usually safer than pure English.
            ocr_lang = "chi_sim+eng"

    tuned_ocr_lang, tuned_borderless, tuned_confidence, tuned_implicit_rows, tuned_implicit_columns = tune_ocr_options(
        ocr_lang,
        args.borderless,
        args.img2table_min_confidence,
        ocr_auto_tune,
    )
    if ocr_auto_tune:
        log(
            (
                "Auto OCR tuning enabled: "
                f"lang={tuned_ocr_lang}, borderless={tuned_borderless}, "
                f"min_confidence={tuned_confidence}, implicit_rows={tuned_implicit_rows}, "
                f"implicit_columns={tuned_implicit_columns}"
            ),
            args.verbose,
        )

    extracted: List[ExtractedTable] = []

    if args.mode == "camelot":
        extracted.extend(
            extract_with_camelot(
                input_pdf,
                pages,
                args.prefer,
                args.accuracy_threshold,
                args.min_rows,
                args.min_cols,
                args.min_filled_ratio,
                args.row_compact,
                args.row_compact_empty_ratio,
                args.row_compact_header_rows,
                args.verbose,
            )
        )
    elif args.mode == "pdfplumber":
        extracted.extend(
            extract_with_pdfplumber(
                input_pdf,
                pages,
                args.min_rows,
                args.min_cols,
                args.min_filled_ratio,
                args.row_compact,
                args.row_compact_empty_ratio,
                args.row_compact_header_rows,
                args.verbose,
            )
        )
    elif args.mode == "img2table":
        extracted.extend(
            extract_with_img2table(
                input_pdf,
                pages,
                tuned_ocr_lang,
                tuned_borderless,
                tuned_confidence,
                tuned_implicit_rows,
                tuned_implicit_columns,
                args.min_rows,
                args.min_cols,
                args.min_filled_ratio,
                args.row_compact,
                args.row_compact_empty_ratio,
                args.row_compact_header_rows,
                args.verbose,
            )
        )
    else:
        if pdf_kind == "text":
            extracted.extend(
                extract_with_camelot(
                    input_pdf,
                    pages,
                    args.prefer,
                    args.accuracy_threshold,
                    args.min_rows,
                    args.min_cols,
                    args.min_filled_ratio,
                    args.row_compact,
                    args.row_compact_empty_ratio,
                    args.row_compact_header_rows,
                    args.verbose,
                )
            )
            extracted.extend(
                extract_with_pdfplumber(
                    input_pdf,
                    pages,
                    args.min_rows,
                    args.min_cols,
                    args.min_filled_ratio,
                    args.row_compact,
                    args.row_compact_empty_ratio,
                    args.row_compact_header_rows,
                    args.verbose,
                )
            )
            if not extracted:
                log("No text-based tables found; trying OCR fallback (img2table).", args.verbose)
                extracted.extend(
                    extract_with_img2table(
                        input_pdf,
                        pages,
                        tuned_ocr_lang,
                        tuned_borderless,
                        tuned_confidence,
                        tuned_implicit_rows,
                        tuned_implicit_columns,
                        args.min_rows,
                        args.min_cols,
                        args.min_filled_ratio,
                        args.row_compact,
                        args.row_compact_empty_ratio,
                        args.row_compact_header_rows,
                        args.verbose,
                    )
                )
        else:
            extracted.extend(
                extract_with_img2table(
                    input_pdf,
                    pages,
                    tuned_ocr_lang,
                    tuned_borderless,
                    tuned_confidence,
                    tuned_implicit_rows,
                    tuned_implicit_columns,
                    args.min_rows,
                    args.min_cols,
                    args.min_filled_ratio,
                    args.row_compact,
                    args.row_compact_empty_ratio,
                    args.row_compact_header_rows,
                    args.verbose,
                )
            )
            extracted.extend(
                extract_with_pdfplumber(
                    input_pdf,
                    pages,
                    args.min_rows,
                    args.min_cols,
                    args.min_filled_ratio,
                    args.row_compact,
                    args.row_compact_empty_ratio,
                    args.row_compact_header_rows,
                    args.verbose,
                    )
                )

    deduped = deduplicate_tables(extracted)
    return select_best_table_per_page(deduped, verbose=args.verbose)


def build_output_path(args: argparse.Namespace, input_pdf: Path, batch_mode: bool) -> Path:
    if not batch_mode and args.output is not None:
        return args.output

    if batch_mode:
        output_dir = args.output_dir or args.input_path / "extracted_tables"
        return output_dir / f"{input_pdf.stem}_tables.xlsx"

    if args.output_dir is not None:
        return args.output_dir / f"{input_pdf.stem}_tables.xlsx"

    return input_pdf.with_name(f"{input_pdf.stem}_tables.xlsx")


def is_monthly_report_file(input_pdf: Path) -> bool:
    return "广东电力现货市场结算运行情况月报" in input_pdf.name


def is_shandong_monthly_report_file(input_pdf: Path) -> bool:
    return "山东电力市场信息披露月报" in input_pdf.name


def write_shandong_excel(
    output_path: Path,
    shandong_result: ShandongExtractionResult,
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    info_df = pd.DataFrame(
        shandong_result.info_rows,
        columns=["报告月份", "section", "subsection", "field", "value", "unit", "source_text", "notes"],
    )
    if info_df.empty:
        info_df = pd.DataFrame(
            columns=["报告月份", "section", "subsection", "field", "value", "unit", "source_text", "notes"]
        )
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        info_df.to_excel(writer, index=False, sheet_name="山东_信息汇总")
        for sheet_name in ["山东_表2_中长期交易情况_raw", "山东_表3_现货交易情况_raw", "山东_表8_市场运行费用_raw"]:
            table_df = shandong_result.raw_tables.get(sheet_name, pd.DataFrame(columns=["raw"]))
            table_df.to_excel(writer, index=False, sheet_name=sheet_name[:31])
        diag_df = pd.DataFrame({"diagnostics": shandong_result.diagnostics})
        diag_df.to_excel(writer, index=False, sheet_name="_diagnostics")


def main() -> int:
    args = parse_args()

    if args.output is not None and args.output_dir is not None:
        print("Use either --output or --output-dir, not both.", file=sys.stderr)
        return 1

    try:
        input_pdfs = collect_input_pdfs(args.input_path)
    except ValueError as exc:
        print(str(exc), file=sys.stderr)
        return 1

    if not input_pdfs:
        print(f"No PDF files found in: {args.input_path}", file=sys.stderr)
        return 1

    if len(input_pdfs) > 1 and args.output is not None:
        print("--output can only be used with a single PDF input.", file=sys.stderr)
        return 1

    batch_mode = len(input_pdfs) > 1 or args.input_path.is_dir()
    failures = 0

    for input_pdf in input_pdfs:
        try:
            if args.demo_section_metrics:
                demo_sections = demo_extract_market_section_metrics()
                print("[DEMO] Section metric extraction rows:")
                for section in demo_sections:
                    print(f" [{section.section_title}]")
                    for title_month_date, name, value, unit, report_date in section.rows:
                        print(
                            f" - title_month_date={title_month_date}, metric={name}, value={value}, unit={unit}, date={report_date}"
                        )
                return 0
            monthly_report = is_monthly_report_file(input_pdf)
            shandong_report = is_shandong_monthly_report_file(input_pdf)
            extracted: List[ExtractedTable] = []
            if monthly_report:
                attach1_border_tables = extract_attach1_with_border_grid(input_pdf, args.verbose)
                if attach1_border_tables:
                    extracted = attach1_border_tables

            if not extracted:
                extracted = extract_tables_for_pdf(input_pdf, args)
            if not extracted:
                print(
                    (
                        f"No tables extracted from {input_pdf.name}. If scanned, install img2table and OCR support, "
                        "then retry with --mode img2table --ocr-lang-auto --verbose."
                    ),
                    file=sys.stderr,
                )
                failures += 1
                continue
            if monthly_report:
                extracted = postprocess_tables_for_monthly_report(extracted)
                extracted = select_attach1_tables_for_monthly_report(extracted)
                section_results = extract_configured_sections_from_pdf(
                    str(input_pdf),
                    default_section_configs(),
                )
            elif shandong_report:
                section_results = []
            else:
                section_results = extract_configured_sections_from_pdf(
                    str(input_pdf),
                    default_section_configs(),
                )
            output_path = build_output_path(args, input_pdf, batch_mode)
            if shandong_report:
                shandong_result = extract_shandong_market_disclosure_monthly_report(
                    pdf_path=str(input_pdf),
                    text=None,
                    tables=extracted,
                    output_path=str(output_path),
                    diagnostics=[],
                )
                write_shandong_excel(output_path, shandong_result)
                if args.verbose:
                    for diag in shandong_result.diagnostics:
                        log(f"[Shandong] {diag}", args.verbose)
            else:
                write_excel(
                    output_path,
                    extracted,
                    excel_style_mode=args.excel_style_mode,
                    section_results=section_results,
                    include_summary_sheet=not monthly_report,
                    table_sheet_base_name="附表1" if monthly_report else "Table",
                    table_write_header=not monthly_report,
                )
            print(f"[OK] {input_pdf.name}: saved {len(extracted)} table(s) to {output_path}")
        except Exception as exc:
            print(f"[FAILED] {input_pdf}: {exc}", file=sys.stderr)
            failures += 1

    if failures == len(input_pdfs):
        return 2
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
