import unittest

from openpyxl import Workbook

from pdf_table_extractor import overlay_shandong_manual_frameworks
from shandong_monthly_extractor import ShandongExtractionResult
import pandas as pd


class TestShandongFrameworkOverlay(unittest.TestCase):
    def test_overlay_table2_table3_and_table8(self):
        wb = Workbook()
        wb.remove(wb.active)
        s2 = wb.create_sheet("山东_表2_中长期交易情况")
        s3 = wb.create_sheet("山东_表3_现货交易情况")
        s8 = wb.create_sheet("山东_表8_市场运行费用_总体结算情况")
        s8["C20"] = "raw-keep"

        result = ShandongExtractionResult(
            info_rows=[],
            raw_tables={
                "山东_表2_中长期交易情况": pd.DataFrame(),
                "山东_表3_现货交易情况": pd.DataFrame(),
                "山东_表8_市场运行费用": pd.DataFrame(),
            },
            diagnostics=[],
            report_month="2025-03",
        )

        overlay_shandong_manual_frameworks(wb, result)

        self.assertEqual(s2["A1"].value, "单位：亿千瓦时、元/兆瓦时")
        self.assertEqual(s2["B3"].value, "月度累计交易电量")
        self.assertEqual(s2["C3"].value, "加权平均电价")

        self.assertEqual(s3["A2"].value, "日期")
        self.assertEqual(s3["B2"].value, "发电侧日前出清电量")
        self.assertEqual(s3["A3"].value.strftime("%Y-%m-%d"), "2025-03-01")
        self.assertEqual(s3["A33"].value.strftime("%Y-%m-%d"), "2025-03-31")

        self.assertEqual(s8["A1"].value, "单位：万元、元/兆瓦时")
        self.assertEqual(s8["A2"].value, "序号")
        self.assertEqual(s8["B2"].value, "类别")
        self.assertEqual(s8["C2"].value, "费用总额")
        self.assertEqual(s8["D2"].value, "分摊返还均价")
        self.assertEqual(s8["E2"].value, "分摊返还主体")
        self.assertEqual(s8["A3"].value, 1)
        self.assertEqual(s8["A14"].value, 12)
        self.assertEqual(s8["B3"].value, "启动费用")
        self.assertEqual(s8["B14"].value, "优发优购曲线匹配偏差费用")
        self.assertEqual(s8["C20"].value, "raw-keep")


if __name__ == "__main__":
    unittest.main()
